import os
import glob
import openpyxl

folder_path = "c:/dev"

for filename in glob.glob(os.path.join(folder_path, '*.xlsx')):

    print(filename)

    wb = openpyxl.open(filename)
    ws = wb.active

    #Update cells
    print('old value:', ws['A1'].value)
    ws['A1'] = 'Hello World'
    print('new value:', ws['A1'].value)

    wb.save(filename)